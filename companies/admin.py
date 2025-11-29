"""
Admin configuration for companies app
"""
from django.contrib import admin
from django.contrib.admin import SimpleListFilter
from django.db import models
from django.utils.html import format_html
from django.urls import reverse
from django.utils.safestring import mark_safe
from django.utils import timezone
from django.http import HttpResponse
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from .models import Company, ActivationSchedule
from .utils import format_riyadh_datetime, format_arabic_datetime


class ActivationStatusFilter(SimpleListFilter):
    """Custom filter for activation status"""
    title = 'حالة التفعيل'
    parameter_name = 'activation_status'
    
    def lookups(self, request, model_admin):
        return (
            ('permanent', 'مفعل بشكل دائم'),
            ('scheduled', 'مفعل حسب الجدولة'),
            ('inactive', 'ملغي التفعيل'),
            ('temporary', 'مفعل مؤقتاً'),
        )
    
    def queryset(self, request, queryset):
        if self.value() == 'permanent':
            return queryset.filter(
                is_active=True,
                activation_start_time__isnull=True,
                activation_end_time__isnull=True
            )
        elif self.value() == 'scheduled':
            return queryset.filter(
                is_active=True,
                schedules__is_active=True
            ).distinct()
        elif self.value() == 'inactive':
            return queryset.filter(is_active=False)
        elif self.value() == 'temporary':
            return queryset.filter(
                is_active=True,
                activation_start_time__isnull=False
            )
        return queryset


class ScheduleStatusFilter(SimpleListFilter):
    """Custom filter for schedule status"""
    title = 'حالة الجدولة'
    parameter_name = 'schedule_status'
    
    def lookups(self, request, model_admin):
        return (
            ('has_schedules', 'لديها جداول'),
            ('no_schedules', 'بدون جداول'),
            ('active_schedules', 'جداول نشطة'),
            ('inactive_schedules', 'جداول متوقفة'),
        )
    
    def queryset(self, request, queryset):
        if self.value() == 'has_schedules':
            return queryset.filter(schedules__isnull=False).distinct()
        elif self.value() == 'no_schedules':
            return queryset.filter(schedules__isnull=True)
        elif self.value() == 'active_schedules':
            return queryset.filter(schedules__is_active=True).distinct()
        elif self.value() == 'inactive_schedules':
            return queryset.filter(schedules__is_active=False).distinct()
        return queryset


class DynamicStatusFilter(SimpleListFilter):
    """Custom filter for dynamic status"""
    title = 'الحالة الفعلية'
    parameter_name = 'dynamic_status'
    
    def lookups(self, request, model_admin):
        return (
            ('active', 'نشط الآن'),
            ('scheduled', 'مجدول'),
            ('inactive', 'غير نشط'),
            ('pending', 'قيد المراجعة'),
            ('approved', 'موافق عليه'),
            ('rejected', 'مرفوض'),
        )
    
    def queryset(self, request, queryset):
        if self.value() == 'active':
            return queryset.filter(is_active=True).filter(
                models.Q(activation_start_time__lte=timezone.now()) &
                models.Q(activation_end_time__gte=timezone.now())
            )
        elif self.value() == 'scheduled':
            return queryset.filter(schedules__is_active=True).distinct()
        elif self.value() == 'inactive':
            return queryset.filter(is_active=False)
        elif self.value() == 'pending':
            return queryset.filter(status='pending')
        elif self.value() == 'approved':
            return queryset.filter(status='approved')
        elif self.value() == 'rejected':
            return queryset.filter(status='rejected')
        return queryset


class CurrentStatusFilter(SimpleListFilter):
    """Custom filter for current activation status"""
    title = 'الحالة الحالية'
    parameter_name = 'current_status'
    
    def lookups(self, request, model_admin):
        return (
            ('currently_active', 'نشط الآن'),
            ('currently_inactive', 'غير نشط الآن'),
            ('expired', 'منتهي الصلاحية'),
            ('upcoming', 'قادم قريباً'),
        )
    
    def queryset(self, request, queryset):
        now = timezone.now()
        
        if self.value() == 'currently_active':
            return queryset.filter(
                is_active=True,
                activation_start_time__lte=now,
                activation_end_time__gte=now
            )
        elif self.value() == 'currently_inactive':
            return queryset.filter(
                models.Q(is_active=False) |
                models.Q(activation_end_time__lt=now) |
                models.Q(activation_start_time__gt=now)
            )
        elif self.value() == 'expired':
            return queryset.filter(
                is_active=True,
                activation_end_time__lt=now
            )
        elif self.value() == 'upcoming':
            return queryset.filter(
                is_active=True,
                activation_start_time__gt=now
            )
        return queryset


class ActivationScheduleInline(admin.TabularInline):
    """Inline admin for activation schedules"""
    model = ActivationSchedule
    extra = 1
    fields = [
        'is_active',
        'saturday', 'sunday', 'monday', 'tuesday', 'wednesday', 'thursday', 'friday',
        'start_hour', 'end_hour', 'duration_hours_display',
        'last_activation'
    ]
    readonly_fields = ['duration_hours_display', 'last_activation']
    
    classes = ['collapse']
    
    verbose_name = "جدولة تفعيل"
    verbose_name_plural = "📅 جدولة التفعيل التلقائي (يتم حساب عدد الساعات تلقائياً)"
    
    def duration_hours_display(self, obj):
        """Display calculated duration"""
        if obj.start_hour is not None and obj.end_hour is not None:
            if obj.start_hour <= obj.end_hour:
                duration = obj.end_hour - obj.start_hour
                if duration == 0:
                    duration = 1
            else:
                duration = (24 - obj.start_hour) + obj.end_hour
            return f"{duration} ساعة (محسوبة تلقائياً)"
        return "-"
    duration_hours_display.short_description = "المدة"


@admin.register(Company)
class CompanyAdmin(admin.ModelAdmin):
    inlines = [ActivationScheduleInline]
    
    class Media:
        js = ('admin/js/company_status_updater.js', 'admin/js/prize_percentages.js',)
        css = {
            'all': ('admin/css/prize_percentages.css',)
        }
    
    def save_model(self, request, obj, form, change):
        """Override save to handle prize percentages"""
        # Save the model first
        super().save_model(request, obj, form, change)
        
        # Get prizes
        prizes = obj.get_prizes_list()
        if not prizes:
            return
        
        # Collect percentages from form data (POST request)
        prize_percentages = []
        for i in range(len(prizes)):
            field_name = f'prize_percentage_{i}'
            if field_name in request.POST:
                try:
                    percentage = int(request.POST[field_name])
                    if percentage > 0:  # Only accept positive values
                        prize_percentages.append(percentage)
                except (ValueError, TypeError):
                    pass
        
        # If percentages were submitted and match prizes count, normalize and save them
        if prize_percentages and len(prize_percentages) == len(prizes):
            # Normalize percentages to sum to 100
            total = sum(prize_percentages)
            if total > 0:
                # Normalize: (each_percentage / total) * 100
                normalized = [(float(p) / total) * 100.0 for p in prize_percentages]
                prize_percentages = [round(p) for p in normalized]
                
                # Adjust to ensure sum is exactly 100 (handle rounding errors)
                current_sum = sum(prize_percentages)
                if current_sum != 100:
                    difference = 100 - current_sum
                    max_idx = prize_percentages.index(max(prize_percentages))
                    prize_percentages[max_idx] += difference
                
                # Ensure no percentage is less than 1
                for i in range(len(prize_percentages)):
                    if prize_percentages[i] < 1:
                        prize_percentages[i] = 1
                
                # Re-adjust if needed
                current_sum = sum(prize_percentages)
                if current_sum != 100:
                    difference = 100 - current_sum
                    max_idx = prize_percentages.index(max(prize_percentages))
                    prize_percentages[max_idx] += difference
            
            # Store in notes field
            prizes_with_percentages = [
                {'name': prize, 'percentage': percentage}
                for prize, percentage in zip(prizes, prize_percentages)
            ]
            obj.notes = json.dumps({
                'prize_percentages': prize_percentages,
                'prizes_with_percentages': prizes_with_percentages
            }, ensure_ascii=False)
            obj.save(update_fields=['notes'])
    
    list_display = [
        'name', 
        'slug',
        'final_type', 
        'email', 
        'dynamic_status_display',
        'is_active', 
        'activation_status_display',
        'activation_type_display',
        'calculated_active_hours_display',
        'has_schedules',
        'created_at',
        'company_link'
    ]
    list_filter = ['status', 'is_active', 'type', 'created_at', 'updated_at', 'active_hours', DynamicStatusFilter, ActivationStatusFilter, ScheduleStatusFilter, CurrentStatusFilter]
    search_fields = ['name', 'slug', 'email', 'phone', 'type', 'custom_type']
    readonly_fields = ['slug', 'created_at', 'updated_at', 'approved_at', 'company_link', 'activation_status', 'schedules_summary', 'prize_percentages_editor']
    
    fieldsets = (
        ('المعلومات الأساسية', {
            'fields': ('name', 'slug', 'type', 'custom_type', 'email', 'phone')
        }),
        ('إعدادات اللعبة', {
            'fields': ('prizes', 'colors', 'logo_url', 'prize_percentages_editor')
        }),
        ('الحالة والإدارة', {
            'fields': ('status', 'is_active', 'active_hours', 'activation_start_time', 'activation_end_time', 'activation_status', 'notes')
        }),
        ('📅 ملخص الجدولة التلقائية', {
            'fields': ('schedules_summary',),
            'description': 'يمكنك إضافة جداول التفعيل التلقائي في الأسفل'
        }),
        ('التواريخ', {
            'fields': ('created_at', 'updated_at', 'approved_at'),
            'classes': ('collapse',)
        }),
        ('رابط اللعبة', {
            'fields': ('company_link',),
        }),
    )
    
    actions = ['activate_companies', 'deactivate_companies', 'activate_by_schedule', 'export_to_excel', 'delete_selected']
    
    def final_type(self, obj):
        return obj.final_type
    final_type.short_description = 'نوع الجهة'
    
    def company_link(self, obj):
        if obj.slug:
            url = reverse('game:play', kwargs={'slug': obj.slug})
            full_url = f'{url}'
            return format_html(
                '<a href="{}" target="_blank" style="color: #6A3FA0; font-weight: bold;">🎡 {}</a><br>'
                '<code style="background: #f0f0f0; padding: 5px; border-radius: 3px; font-size: 11px;">{}</code>',
                url, obj.name, full_url
            )
        return '-'
    company_link.short_description = 'رابط اللعبة'
    
    def activation_status(self, obj):
        """عرض حالة التفعيل بتوقيت السعودية مع توضيح النوع"""
        if not obj.is_active:
            return format_html('<span style="color: #999; font-weight: bold;">⭕ غير مفعّل</span>')
        
        if obj.is_currently_active:
            if obj.activation_end_time:
                # تفعيل مؤقت محدد
                formatted_time = format_arabic_datetime(obj.activation_end_time)
                
                return format_html(
                    '<div style="background: #e8f5e9; padding: 8px; border-radius: 6px; border-left: 4px solid #4caf50;">'
                    '<span style="color: #2e7d32; font-weight: bold;">⏰ تفعيل مؤقت محدد</span><br>'
                    '<span style="color: #333; font-size: 12px;">ينتهي: {}</span><br>'
                    '<small style="color: #6A3FA0; font-weight: 600;">🕐 توقيت السعودية</small>'
                    '</div>',
                    formatted_time
                )
            else:
                # تفعيل دائم مستمر
                return format_html(
                    '<div style="background: #e3f2fd; padding: 8px; border-radius: 6px; border-left: 4px solid #2196f3;">'
                    '<span style="color: #1565c0; font-weight: bold;">♾️ تفعيل دائم مستمر</span><br>'
                    '<small style="color: #666;">بدون حد زمني</small>'
                    '</div>'
                )
        else:
            if obj.activation_end_time:
                formatted_time = format_arabic_datetime(obj.activation_end_time)
                return format_html(
                    '<div style="background: #fff3e0; padding: 8px; border-radius: 6px; border-left: 4px solid #ff9800;">'
                    '<span style="color: #e65100; font-weight: bold;">⏸️ منتهي التفعيل</span><br>'
                    '<small style="color: #999;">انتهى في: {}</small>'
                    '</div>',
                    formatted_time
                )
            return format_html(
                '<span style="color: orange; font-weight: bold;">⏸️ منتهي التفعيل</span>'
            )
    activation_status.short_description = 'حالة التفعيل'
    
    def has_schedules(self, obj):
        """Show if company has active schedules"""
        schedules = obj.schedules.filter(is_active=True)
        count = schedules.count()
        if count > 0:
            return format_html(
                '<span style="color: #28a745; font-weight: bold;">✓ {} جدولة</span>',
                count
            )
        return format_html('<span style="color: #999;">-</span>')
    has_schedules.short_description = 'جدولة تلقائية'
    
    def activation_status_display(self, obj):
        """Display activation status with colors"""
        status = obj.activation_status_display
        is_active = obj.is_currently_active
        
        if is_active:
            if "دائم" in status:
                return format_html('<span style="color: #28a745; font-weight: bold;">✅ {}</span>', status)
            else:
                return format_html('<span style="color: #17a2b8; font-weight: bold;">⏰ {}</span>', status)
        else:
            return format_html('<span style="color: #dc3545; font-weight: bold;">❌ {}</span>', status)
    activation_status_display.short_description = 'حالة التفعيل المباشرة'
    
    def activation_type_display(self, obj):
        """Display activation type"""
        if not obj.is_active:
            return format_html('<span style="color: #dc3545; font-weight: bold;">❌ ملغي التفعيل</span>')
        
        # Check if it has active schedules
        has_active_schedules = obj.schedules.filter(is_active=True).exists()
        
        # Check if it's permanently active (no start/end time)
        is_permanent = not obj.activation_start_time and not obj.activation_end_time
        
        if is_permanent:
            return format_html('<span style="color: #28a745; font-weight: bold;">🔄 مفعل بشكل دائم</span>')
        elif has_active_schedules:
            return format_html('<span style="color: #17a2b8; font-weight: bold;">📅 مفعل حسب الجدولة</span>')
        elif obj.activation_start_time:
            return format_html('<span style="color: #ffc107; font-weight: bold;">⏰ مفعل مؤقتاً</span>')
        else:
            return format_html('<span style="color: #6c757d; font-weight: bold;">❓ غير محدد</span>')
    activation_type_display.short_description = 'نوع التفعيل'
    
    def dynamic_status_display(self, obj):
        """Display dynamic status with colors"""
        status = obj.dynamic_status
        
        if status == 'active':
            return format_html('<span style="color: #28a745; font-weight: bold;">🟢 نشط الآن</span>')
        elif status == 'scheduled':
            return format_html('<span style="color: #17a2b8; font-weight: bold;">📅 مجدول</span>')
        elif status == 'inactive':
            return format_html('<span style="color: #dc3545; font-weight: bold;">🔴 غير نشط</span>')
        elif status == 'pending':
            return format_html('<span style="color: #ffc107; font-weight: bold;">⏳ قيد المراجعة</span>')
        elif status == 'approved':
            return format_html('<span style="color: #28a745; font-weight: bold;">✅ موافق عليه</span>')
        elif status == 'rejected':
            return format_html('<span style="color: #dc3545; font-weight: bold;">❌ مرفوض</span>')
        else:
            return format_html('<span style="color: #6c757d; font-weight: bold;">❓ غير محدد</span>')
    dynamic_status_display.short_description = 'الحالة الفعلية'
    
    def calculated_active_hours_display(self, obj):
        """Display calculated active hours - show schedule hours if company has schedules"""
        # If company has active schedules, show schedule duration
        active_schedules = obj.schedules.filter(is_active=True)
        if active_schedules.exists():
            # Get first active schedule's duration
            schedule_hours = active_schedules.first().duration_hours
            return format_html('<span style="color: #17a2b8; font-weight: bold;">📅 {} ساعة (من الجدولة)</span>', schedule_hours)
        
        # Otherwise, show calculated hours
        hours = obj.calculated_active_hours
        
        if hours == 0:
            return format_html('<span style="color: #dc3545; font-weight: bold;">❌ 0 ساعة</span>')
        elif hours == 24:
            return format_html('<span style="color: #28a745; font-weight: bold;">🔄 24 ساعة (دائم)</span>')
        elif hours < 24:
            return format_html('<span style="color: #17a2b8; font-weight: bold;">⏰ {} ساعة</span>', hours)
        else:
            days = hours // 24
            remaining_hours = hours % 24
            if remaining_hours > 0:
                return format_html('<span style="color: #6c757d; font-weight: bold;">📅 {} يوم و {} ساعة</span>', days, remaining_hours)
            else:
                return format_html('<span style="color: #6c757d; font-weight: bold;">📅 {} يوم</span>', days)
    calculated_active_hours_display.short_description = 'عدد ساعات التفعيل'
    
    def prize_percentages_editor(self, obj):
        """Display prize percentages in editable fields"""
        if not obj.pk:
            return format_html('<p style="color: #999; padding: 15px; background: #f8f9fa; border-radius: 5px;">⚠️ احفظ الشركة أولاً لعرض وتعديل النسب المئوية</p>')
        
        prizes = obj.get_prizes_list()
        if not prizes:
            return format_html('<p style="color: #999; padding: 15px; background: #f8f9fa; border-radius: 5px;">⚠️ لا توجد جوائز. أضف جوائز أولاً.</p>')
        
        # Get percentages from notes
        prize_percentages = []
        if obj.notes:
            try:
                notes_data = json.loads(obj.notes)
                if 'prize_percentages' in notes_data:
                    prize_percentages = notes_data['prize_percentages']
            except (json.JSONDecodeError, KeyError):
                pass
        
        # Ensure percentages match prizes count
        if len(prize_percentages) != len(prizes):
            equal_percentage = 100 // len(prizes)
            prize_percentages = [equal_percentage] * len(prizes)
            remainder = 100 - (equal_percentage * len(prizes))
            if remainder > 0:
                prize_percentages[-1] += remainder
        
        # Calculate total
        total = sum(prize_percentages)
        total_color = '#28a745' if total == 100 else '#ffc107' if total < 100 else '#dc3545'
        
        html = f'''
        <div style="padding: 15px; background: #f8f9fa; border-radius: 5px; margin: 10px 0; border: 2px solid #6A3FA0;">
            <h3 style="margin-top: 0; color: #6A3FA0; font-size: 18px;">🎯 النسب المئوية للجوائز</h3>
            <p style="color: #666; font-size: 13px; margin-bottom: 15px; background: #e3f2fd; padding: 10px; border-radius: 5px;">
                💡 <strong>كيف تعمل النسب:</strong> كلما زادت النسبة زاد احتمال الفوز بالجائزة. يمكنك إدخال أي قيمة (مثلاً 300%) وسيتم تطبيع النسب تلقائياً إلى 100%.
            </p>
            <table style="width: 100%; border-collapse: collapse; background: white; border-radius: 5px; overflow: hidden;">
                <thead>
                    <tr style="background: #6A3FA0; color: white;">
                        <th style="padding: 12px; text-align: right; border: 1px solid #ddd; font-weight: bold;">الجائزة</th>
                        <th style="padding: 12px; text-align: center; border: 1px solid #ddd; width: 200px; font-weight: bold;">النسبة المئوية</th>
                    </tr>
                </thead>
                <tbody>
        '''
        
        for i, (prize, percentage) in enumerate(zip(prizes, prize_percentages)):
            field_name = f'prize_percentage_{i}'
            html += f'''
                    <tr>
                        <td style="padding: 12px; border: 1px solid #ddd; background: white;">
                            <strong style="color: #333;">{prize}</strong>
                        </td>
                        <td style="padding: 12px; border: 1px solid #ddd; background: white; text-align: center;">
                            <div style="display: flex; align-items: center; justify-content: center; gap: 8px;">
                                <input type="number" 
                                       id="{field_name}"
                                       name="{field_name}" 
                                       value="{percentage}" 
                                       min="1" 
                                       max="100" 
                                       required
                                       class="prize-percentage-input"
                                       style="width: 100px; padding: 8px; text-align: center; border: 2px solid #6A3FA0; border-radius: 5px; font-weight: bold; font-size: 14px;"
                                       onchange="updateTotalPercentage()"
                                       oninput="updateTotalPercentage()">
                                <span style="color: #6A3FA0; font-weight: bold; font-size: 16px;">%</span>
                            </div>
                        </td>
                    </tr>
            '''
        
        html += f'''
                </tbody>
                <tfoot>
                    <tr style="background: #e3f2fd;">
                        <td style="padding: 12px; border: 1px solid #ddd; text-align: right; font-weight: bold; font-size: 16px;">
                            المجموع:
                        </td>
                        <td style="padding: 12px; border: 1px solid #ddd; text-align: center;">
                            <span id="total-percentage" style="color: {total_color}; font-weight: bold; font-size: 18px;">
                                {total}%
                            </span>
                            <small style="display: block; color: #666; margin-top: 5px; font-size: 12px;">
                                (سيتم تطبيع النسب تلقائياً عند الحفظ)
                            </small>
                        </td>
                    </tr>
                </tfoot>
            </table>
        </div>
        '''
        
        return format_html(html)
    prize_percentages_editor.short_description = '🎯 تعديل النسب المئوية للجوائز'
    prize_percentages_editor.allow_tags = True
    
    def schedules_summary(self, obj):
        """Display summary of active schedules"""
        schedules = obj.schedules.all()
        
        if not schedules.exists():
            return format_html(
                '<div style="padding: 15px; background: #f8f9fa; border-radius: 5px; border-right: 4px solid #ffc107;">'
                '<p style="margin: 0; color: #856404;">⚠️ لا توجد جداول تفعيل مضافة بعد</p>'
                '<small style="color: #666;">يمكنك إضافة جداول في الأسفل لتفعيل العجلة تلقائياً</small>'
                '</div>'
            )
        
        html = '<div style="padding: 10px; background: #f8f9fa; border-radius: 5px;">'
        
        for schedule in schedules:
            status_color = '#28a745' if schedule.is_active else '#dc3545'
            status_icon = '✅' if schedule.is_active else '⏸️'
            status_text = 'مفعلة' if schedule.is_active else 'متوقفة'
            
            active_now = schedule.should_activate_now() if schedule.is_active else False
            now_badge = '<span style="background: #28a745; color: white; padding: 2px 8px; border-radius: 10px; font-size: 11px; margin-right: 5px;">نشطة الآن</span>' if active_now else ''
            
            html += f'''
            <div style="margin-bottom: 10px; padding: 10px; background: white; border-radius: 5px; border-right: 4px solid {status_color};">
                <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 5px;">
                    <strong style="color: {status_color};">{status_icon} {status_text}</strong>
                    {now_badge}
                </div>
                <div style="font-size: 13px; color: #666;">
                    <p style="margin: 3px 0;"><strong>الأيام:</strong> {schedule.get_active_days_display()}</p>
                    <p style="margin: 3px 0;"><strong>الوقت:</strong> {schedule.get_time_display()}</p>
                    <p style="margin: 3px 0;"><strong>المدة:</strong> {schedule.duration_hours} ساعة</p>
                </div>
            </div>
            '''
        
        html += '</div>'
        
        return format_html(html)
    schedules_summary.short_description = 'ملخص الجدولة'
    
    def activate_companies(self, request, queryset):
        """تفعيل الشركات المحددة بشكل دائم (بدون حد زمني)"""
        updated = queryset.update(
            is_active=True, 
            status='approved',
            activation_start_time=None,
            activation_end_time=None
        )
        self.message_user(
            request, 
            f'✅ تم تفعيل {updated} شركة بشكل دائم (تفعيل مستمر بدون حد زمني).',
            level='success'
        )
    activate_companies.short_description = '✅ تفعيل دائم (مستمر بدون حد زمني)'
    
    def deactivate_companies(self, request, queryset):
        """إلغاء تفعيل الشركات المحددة"""
        updated = queryset.update(
            is_active=False,
            activation_start_time=None,
            activation_end_time=None
        )
        self.message_user(
            request, 
            f'تم إلغاء تفعيل {updated} شركة.',
            level='warning'
        )
    deactivate_companies.short_description = '❌ إلغاء تفعيل الشركات المحددة'
    
    
    def activate_by_schedule(self, request, queryset):
        """تفعيل الشركات حسب جداولها الحالية"""
        activated_count = 0
        exact_hour_count = 0
        no_schedule_count = 0
        details = []
        exact_hour_details = []
        
        for company in queryset:
            # Get active schedules
            active_schedules = company.schedules.filter(is_active=True)
            
            if not active_schedules.exists():
                no_schedule_count += 1
                continue
            
            # Try to activate from any matching schedule
            activated = False
            for schedule in active_schedules:
                can_activate, is_exact_hour, message = schedule.can_activate_manually()
                
                if is_exact_hour:
                    # Exactly at start_hour - show message only
                    exact_hour_count += 1
                    exact_hour_details.append(f"⏰ {company.name}: {message}")
                    activated = True
                    break
                elif can_activate:
                    # Can activate immediately (before start_hour by 1 minute or after)
                    company.activate_now(hours=schedule.duration_hours, scheduled_hour=schedule.start_hour, scheduled_end_hour=schedule.end_hour)
                    schedule.last_activation = timezone.now()
                    schedule.save()
                    
                    activated_count += 1
                    end_time = format_arabic_datetime(company.activation_end_time)
                    details.append(f"✅ {company.name}: تم التفعيل لـ {schedule.duration_hours} ساعة (حتى {end_time})")
                    activated = True
                    break
        
        # Build message - only show activated and exact hour messages
        message_parts = []
        
        if activated_count > 0:
            message_parts.append(f'✅ تم تفعيل {activated_count} شركة حسب جداولها')
        
        if exact_hour_count > 0:
            message_parts.append(f'⏰ {exact_hour_count} شركة: الوقت الحالي هو نفس وقت بداية الجدولة')
            if exact_hour_details:
                message_parts.append('\nالتفاصيل:')
                message_parts.extend(exact_hour_details[:10])
                if len(exact_hour_details) > 10:
                    message_parts.append(f'... و {len(exact_hour_details) - 10} شركة أخرى')
        
        if no_schedule_count > 0:
            message_parts.append(f'⚠️ {no_schedule_count} شركة بدون جداول نشطة')
        
        # Only show details for activated companies
        if details:
            if activated_count > 0:
                message_parts.append('\nشركات تم تفعيلها:')
                message_parts.extend(details[:10])
                if len(details) > 10:
                    message_parts.append(f'... و {len(details) - 10} شركة أخرى')
        
        message = '\n'.join(message_parts) if message_parts else 'لا توجد شركات للتحديث'
        
        level = 'success' if activated_count > 0 else ('info' if exact_hour_count > 0 else 'warning')
        self.message_user(request, message, level=level)
    
    activate_by_schedule.short_description = '📅 تفعيل حسب الجدولة (يتحقق من الأيام والأوقات)'
    
    def changelist_view(self, request, extra_context=None):
        """Override changelist to add export button"""
        # Check if this is an export request
        if 'action' in request.POST and request.POST['action'] == 'export_to_excel':
            # Get all items (or selected items)
            selected_ids = request.POST.getlist('_selected_action')
            if selected_ids:
                queryset = self.get_queryset(request).filter(pk__in=selected_ids)
            else:
                # Export all if nothing selected
                queryset = self.get_queryset(request)
            return self.export_to_excel(request, queryset)
        
        extra_context = extra_context or {}
        extra_context['show_export_button'] = True
        extra_context['export_action_name'] = 'export_to_excel'
        return super().changelist_view(request, extra_context)
    
    def export_to_excel(self, request, queryset):
        """Export companies to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "الشركات"
        
        # Define headers in Arabic
        headers = [
            'ID',
            'الاسم',
            'الاسم المختصر (Slug)',
            'النوع',
            'البريد الإلكتروني',
            'رقم الجوال',
            'الحالة',
            'مفعل',
            'نوع التفعيل',
            'عدد ساعات التفعيل',
            'وقت بداية التفعيل',
            'وقت نهاية التفعيل',
            'عدد الجداول',
            'عدد الجوائز',
            'تاريخ الإنشاء',
            'تاريخ التحديث',
            'تاريخ الموافقة'
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="6A3FA0", end_color="6A3FA0", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        row_num = 2
        for company in queryset.prefetch_related('schedules'):
            prizes = company.get_prizes_list()
            
            ws.cell(row=row_num, column=1, value=company.id)
            ws.cell(row=row_num, column=2, value=company.name)
            ws.cell(row=row_num, column=3, value=company.slug)
            ws.cell(row=row_num, column=4, value=company.final_type)
            ws.cell(row=row_num, column=5, value=company.email or '-')
            ws.cell(row=row_num, column=6, value=company.phone or '-')
            ws.cell(row=row_num, column=7, value=company.get_status_display())
            ws.cell(row=row_num, column=8, value='نعم' if company.is_active else 'لا')
            ws.cell(row=row_num, column=9, value=company.activation_type_display().replace('<span style="color: #', '').split('>')[-1].split('<')[0] if hasattr(company, 'activation_type_display') else company.activation_status_display)
            ws.cell(row=row_num, column=10, value=company.calculated_active_hours)
            
            if company.activation_start_time:
                ws.cell(row=row_num, column=11, value=format_arabic_datetime(company.activation_start_time))
            else:
                ws.cell(row=row_num, column=11, value='-')
            
            if company.activation_end_time:
                ws.cell(row=row_num, column=12, value=format_arabic_datetime(company.activation_end_time))
            else:
                ws.cell(row=row_num, column=12, value='-')
            
            ws.cell(row=row_num, column=13, value=company.schedules.count())
            ws.cell(row=row_num, column=14, value=len(prizes))
            
            if company.created_at:
                ws.cell(row=row_num, column=15, value=format_arabic_datetime(company.created_at))
            else:
                ws.cell(row=row_num, column=15, value='-')
            
            if company.updated_at:
                ws.cell(row=row_num, column=16, value=format_arabic_datetime(company.updated_at))
            else:
                ws.cell(row=row_num, column=16, value='-')
            
            if company.approved_at:
                ws.cell(row=row_num, column=17, value=format_arabic_datetime(company.approved_at))
            else:
                ws.cell(row=row_num, column=17, value='-')
            
            row_num += 1
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if row.value:
                        max_length = max(max_length, len(str(row.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'الشركات_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save workbook to response
        wb.save(response)
        return response
    
    export_to_excel.short_description = "📊 تصدير البيانات المحددة إلى Excel"
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related()


@admin.register(ActivationSchedule)
class ActivationScheduleAdmin(admin.ModelAdmin):
    list_display = [
        'company',
        'get_active_days_short',
        'get_time_range',
        'duration_hours',
        'is_active',
        'last_activation',
        'status_indicator'
    ]
    list_filter = [
        'is_active',
        'monday', 'tuesday', 'wednesday', 'thursday', 
        'friday', 'saturday', 'sunday',
        'created_at'
    ]
    search_fields = ['company__name', 'company__email']
    readonly_fields = ['last_activation', 'created_at', 'updated_at', 'schedule_status_display', 'duration_display']
    
    class Media:
        js = ('admin/js/schedule_status_updater.js', 'admin/js/schedule_delete_handler.js',)
    
    fieldsets = (
        ('معلومات الشركة', {
            'fields': ('company',)
        }),
        ('أيام التفعيل (بترتيب الأسبوع)', {
            'fields': (
                ('saturday', 'sunday', 'monday', 'tuesday'),
                ('wednesday', 'thursday', 'friday'),
            ),
            'description': 'اختر الأيام التي تريد تفعيل الشركة فيها تلقائياً (السبت إلى الجمعة)'
        }),
        ('إعدادات الوقت (نظام 12 ساعة)', {
            'fields': (
                ('start_hour', 'end_hour'),
                'duration_display',
            ),
            'description': '''
                <strong>نظام 12 ساعة - عدد الساعات يحسب تلقائياً:</strong><br>
                <strong>أمثلة:</strong><br>
                • 09:00 صباحاً إلى 05:00 مساءً = 8 ساعات تلقائياً<br>
                • 01:00 ظهراً إلى 10:00 مساءً = 9 ساعات تلقائياً<br>
                • 06:00 مساءً إلى 11:00 مساءً = 5 ساعات تلقائياً
            '''
        }),
        ('الإعدادات', {
            'fields': ('is_active',)
        }),
        ('معلومات التتبع', {
            'fields': ('last_activation', 'schedule_status_display', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def get_active_days_short(self, obj):
        """Get short display of active days (in Arabic week order)"""
        days_short = []
        if obj.saturday: days_short.append('السبت')
        if obj.sunday: days_short.append('الأحد')
        if obj.monday: days_short.append('الاثنين')
        if obj.tuesday: days_short.append('الثلاثاء')
        if obj.wednesday: days_short.append('الأربعاء')
        if obj.thursday: days_short.append('الخميس')
        if obj.friday: days_short.append('الجمعة')
        return ', '.join(days_short) if days_short else '-'
    get_active_days_short.short_description = 'الأيام المفعلة'
    
    def duration_display(self, obj):
        """Display auto-calculated duration"""
        if obj.start_hour <= obj.end_hour:
            duration = obj.end_hour - obj.start_hour
            if duration == 0:
                duration = 1
        else:
            duration = (24 - obj.start_hour) + obj.end_hour
        return format_html(
            '<span style="background: #e3f2fd; padding: 5px 10px; border-radius: 5px; color: #1565c0; font-weight: bold;">'
            '⏱️ {} ساعة (محسوبة تلقائياً)'
            '</span>',
            duration
        )
    duration_display.short_description = 'المدة'
    
    def get_time_range(self, obj):
        """Get time range display"""
        return obj.get_time_display()
    get_time_range.short_description = 'أوقات التفعيل'
    
    def status_indicator(self, obj):
        """Show real-time company activation status"""
        # Get company activation status
        company_status = obj.get_company_activation_status()
        
        # Add data attribute for AJAX updates
        status_id = f"schedule-status-{obj.id}"
        
        if not obj.is_active:
            return format_html(
                '<span id="{}" style="color: #dc3545;">⏸️ الجدولة متوقفة</span>',
                status_id
            )
        
        # Show company activation status
        if company_status['is_active']:
            return format_html(
                '<span id="{}" style="color: {}; font-weight: bold;">✅ {}</span>',
                status_id,
                company_status['color'],
                company_status['display']
            )
        else:
            # Check if should activate soon
            if obj.should_activate_soon():
                return format_html(
                    '<span id="{}" style="color: #ffc107;">⏳ جاهز للتفعيل (ضمن النطاق)</span>',
                    status_id
                )
            else:
                return format_html(
                    '<span id="{}" style="color: #6c757d;">⏳ {} - خارج نطاق الجدولة</span>',
                    status_id,
                    company_status['display']
                )
    status_indicator.short_description = 'حالة الشركة'
    
    def schedule_status_display(self, obj):
        """Detailed schedule status display"""
        now = timezone.now()
        current_day = ['الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة', 'السبت', 'الأحد'][now.weekday()]
        current_time = now.strftime('%H:%M')
        
        status_html = f'''
        <div style="padding: 10px; background: #f8f9fa; border-radius: 5px;">
            <p><strong>الوقت الحالي:</strong> {current_day} - {current_time}</p>
            <p><strong>الأيام المفعلة:</strong> {obj.get_active_days_display()}</p>
            <p><strong>أوقات التفعيل:</strong> {obj.get_time_display()}</p>
            <p><strong>مدة التفعيل:</strong> {obj.duration_hours} ساعة</p>
            <p><strong>حالة الجدولة:</strong> {'مفعلة ✅' if obj.is_active else 'متوقفة ⏸️'}</p>
            <p><strong>يجب التفعيل الآن؟:</strong> {'نعم ✅' if obj.should_activate_now() else 'لا ❌'}</p>
        </div>
        '''
        return format_html(status_html)
    schedule_status_display.short_description = 'حالة الجدولة التفصيلية'
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('company')
    
    actions = ['activate_selected_schedules', 'deactivate_selected_schedules', 'export_to_excel']
    
    def activate_selected_schedules(self, request, queryset):
        """Activate selected schedules"""
        count = queryset.update(is_active=True)
        self.message_user(request, f'تم تفعيل {count} جدولة')
    activate_selected_schedules.short_description = 'تفعيل الجدولة المحددة'
    
    def deactivate_selected_schedules(self, request, queryset):
        """Deactivate selected schedules"""
        count = queryset.update(is_active=False)
        self.message_user(request, f'تم إيقاف {count} جدولة')
    deactivate_selected_schedules.short_description = 'إيقاف الجدولة المحددة'
    
    def changelist_view(self, request, extra_context=None):
        """Override changelist to add export button"""
        # Check if this is an export request
        if 'action' in request.POST and request.POST['action'] == 'export_to_excel':
            # Get all items (or selected items)
            selected_ids = request.POST.getlist('_selected_action')
            if selected_ids:
                queryset = self.get_queryset(request).filter(pk__in=selected_ids)
            else:
                # Export all if nothing selected
                queryset = self.get_queryset(request)
            return self.export_to_excel(request, queryset)
        
        extra_context = extra_context or {}
        extra_context['show_export_button'] = True
        extra_context['export_action_name'] = 'export_to_excel'
        return super().changelist_view(request, extra_context)
    
    def export_to_excel(self, request, queryset):
        """Export activation schedules to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "جداول التفعيل"
        
        # Define headers in Arabic
        headers = [
            'ID',
            'اسم الشركة',
            'البريد الإلكتروني',
            'السبت',
            'الأحد',
            'الاثنين',
            'الثلاثاء',
            'الأربعاء',
            'الخميس',
            'الجمعة',
            'ساعة البداية',
            'ساعة النهاية',
            'المدة (ساعات)',
            'مفعلة',
            'آخر تفعيل',
            'تاريخ الإنشاء',
            'تاريخ التحديث'
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="17a2b8", end_color="17a2b8", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        row_num = 2
        for schedule in queryset.select_related('company'):
            ws.cell(row=row_num, column=1, value=schedule.id)
            ws.cell(row=row_num, column=2, value=schedule.company.name)
            ws.cell(row=row_num, column=3, value=schedule.company.email or '-')
            ws.cell(row=row_num, column=4, value='نعم' if schedule.saturday else 'لا')
            ws.cell(row=row_num, column=5, value='نعم' if schedule.sunday else 'لا')
            ws.cell(row=row_num, column=6, value='نعم' if schedule.monday else 'لا')
            ws.cell(row=row_num, column=7, value='نعم' if schedule.tuesday else 'لا')
            ws.cell(row=row_num, column=8, value='نعم' if schedule.wednesday else 'لا')
            ws.cell(row=row_num, column=9, value='نعم' if schedule.thursday else 'لا')
            ws.cell(row=row_num, column=10, value='نعم' if schedule.friday else 'لا')
            ws.cell(row=row_num, column=11, value=f"{schedule.start_hour}:00")
            ws.cell(row=row_num, column=12, value=f"{schedule.end_hour}:00")
            ws.cell(row=row_num, column=13, value=schedule.duration_hours)
            ws.cell(row=row_num, column=14, value='نعم' if schedule.is_active else 'لا')
            
            if schedule.last_activation:
                ws.cell(row=row_num, column=15, value=format_arabic_datetime(schedule.last_activation))
            else:
                ws.cell(row=row_num, column=15, value='-')
            
            if schedule.created_at:
                ws.cell(row=row_num, column=16, value=format_arabic_datetime(schedule.created_at))
            else:
                ws.cell(row=row_num, column=16, value='-')
            
            if schedule.updated_at:
                ws.cell(row=row_num, column=17, value=format_arabic_datetime(schedule.updated_at))
            else:
                ws.cell(row=row_num, column=17, value='-')
            
            row_num += 1
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if row.value:
                        max_length = max(max_length, len(str(row.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'جداول_التفعيل_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save workbook to response
        wb.save(response)
        return response
    
    export_to_excel.short_description = "📊 تصدير البيانات المحددة إلى Excel"
