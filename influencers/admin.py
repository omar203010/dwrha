"""
Admin configuration for influencers app
"""
from django.contrib import admin
from django.utils.html import format_html
from django.urls import reverse
from django.http import HttpResponse
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from .models import Influencer, Participant


@admin.register(Influencer)
class InfluencerAdmin(admin.ModelAdmin):
    """Admin interface for Influencer model"""
    
    list_display = [
        'id', 'name', 'platform_display', 'username', 'followers_count', 
        'status_display', 'is_active', 'wheel_link_display', 'registration_link_display', 'created_at_display'
    ]
    list_filter = ['status', 'is_active', 'platform', 'created_at']
    search_fields = ['name', 'username', 'email', 'phone']
    readonly_fields = ['slug', 'created_at', 'updated_at', 'approved_at', 'registration_link', 'wheel_link']
    
    fieldsets = (
        ('المعلومات الأساسية', {
            'fields': ('name', 'slug', 'platform', 'custom_platform', 'username', 'profile_url')
        }),
        ('معلومات المنصة', {
            'fields': ('followers_count', 'profile_image_url', 'bio')
        }),
        ('معلومات التواصل', {
            'fields': ('email', 'phone')
        }),
        ('إعدادات اللعبة', {
            'fields': ('prizes', 'colors')
        }),
        ('الحالة والإدارة', {
            'fields': ('status', 'is_active', 'notes')
        }),
        ('التواريخ', {
            'fields': ('created_at', 'updated_at', 'approved_at')
        }),
        ('روابط المؤثر', {
            'fields': ('registration_link', 'wheel_link'),
            'description': '⚠️ تظهر الروابط فقط بعد تفعيل المؤثر (is_active = True)'
        }),
    )
    
    actions = ['approve_influencers', 'reject_influencers', 'activate_influencers', 'deactivate_influencers', 'export_to_excel']
    
    def platform_display(self, obj):
        """Display platform in readable format"""
        if obj.platform == 'other' and obj.custom_platform:
            return obj.custom_platform
        return obj.get_platform_display()
    platform_display.short_description = 'المنصة الرئيسية'
    
    def status_display(self, obj):
        """Display status with emoji"""
        status_dict = {
            'pending': '⏳ قيد المراجعة',
            'approved': '✅ موافق عليه',
            'rejected': '❌ مرفوض',
            'active': '🟢 نشط',
            'inactive': '🔴 غير نشط',
        }
        return status_dict.get(obj.status, obj.get_status_display())
    status_display.short_description = 'الحالة'
    
    def created_at_display(self, obj):
        """Display created at in readable format"""
        if obj.created_at:
            return obj.created_at.strftime('%Y-%m-%d %H:%M')
        return '-'
    created_at_display.short_description = 'تاريخ الإنشاء'
    
    def _get_base_url(self):
        """Get base URL from settings or use relative path"""
        # Use relative URLs - they will work in any domain
        return ''
    
    def registration_link(self, obj):
        """Display registration link for participants"""
        if obj.id and obj.is_active:
            url = reverse('influencers:register_participant', kwargs={'slug': obj.slug})
            return format_html(
                '<div style="margin-bottom: 10px;">'
                '<strong style="display: block; margin-bottom: 5px;">رابط التسجيل للمشاركين:</strong>'
                '<a href="{}" target="_blank" style="color: #6A3FA0; font-weight: bold; word-break: break-all;">{}</a>'
                '</div>',
                url, url
            )
        return format_html('<span style="color: #999; font-style: italic;">⚠️ يظهر بعد تفعيل المؤثر</span>')
    registration_link.short_description = 'رابط التسجيل للمشاركين'
    
    def wheel_link(self, obj):
        """Display wheel game link"""
        if obj.id and obj.is_active:
            url = reverse('influencers:play_wheel', kwargs={'slug': obj.slug})
            return format_html(
                '<div>'
                '<strong style="display: block; margin-bottom: 5px;">رابط العجلة (اختيار الفائز):</strong>'
                '<a href="{}" target="_blank" style="color: #FF6B9D; font-weight: bold; word-break: break-all;">{}</a>'
                '</div>',
                url, url
            )
        return format_html('<span style="color: #999; font-style: italic;">⚠️ يظهر بعد تفعيل المؤثر</span>')
    wheel_link.short_description = 'رابط العجلة'
    
    def wheel_link_display(self, obj):
        """Display wheel link in list view"""
        if obj.id and obj.is_active:
            url = reverse('influencers:play_wheel', kwargs={'slug': obj.slug})
            return format_html(
                '<a href="{}" target="_blank" style="color: #FF6B9D; font-weight: bold;" title="{}">🎡 رابط العجلة</a>',
                url, url
            )
        return format_html('<span style="color: #999;">-</span>')
    wheel_link_display.short_description = 'رابط العجلة'
    
    def registration_link_display(self, obj):
        """Display registration link in list view"""
        if obj.id and obj.is_active:
            url = reverse('influencers:register_participant', kwargs={'slug': obj.slug})
            return format_html(
                '<a href="{}" target="_blank" style="color: #6A3FA0; font-weight: bold;" title="{}">📝 رابط التسجيل</a>',
                url, url
            )
        return format_html('<span style="color: #999;">-</span>')
    registration_link_display.short_description = 'رابط التسجيل'
    
    def changelist_view(self, request, extra_context=None):
        """Override changelist to add export button"""
        # Check if this is an export request
        if 'action' in request.POST and request.POST['action'] == 'export_to_excel':
            # Get all items (or selected items)
            selected_ids = request.POST.getlist('_selected_action')
            if selected_ids:
                queryset = self.get_queryset(request).filter(pk__in=selected_ids)
            else:
                # Export all if nothing selected
                queryset = self.get_queryset(request)
            return self.export_to_excel(request, queryset)
        
        extra_context = extra_context or {}
        extra_context['show_export_button'] = True
        extra_context['export_action_name'] = 'export_to_excel'
        return super().changelist_view(request, extra_context)
    
    def export_to_excel(self, request, queryset):
        """Export influencers to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "المؤثرون"
        
        # Define headers in Arabic
        headers = [
            'ID',
            'الاسم',
            'المنصة الرئيسية',
            'اسم المستخدم',
            'عدد المتابعين',
            'البريد الإلكتروني',
            'رقم الجوال',
            'الحالة',
            'مفعل',
            'عدد الجوائز',
            'عدد المشاركين',
            'تاريخ الإنشاء',
            'تاريخ التحديث',
            'تاريخ الموافقة'
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="FF6B9D", end_color="FF6B9D", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        row_num = 2
        for influencer in queryset.prefetch_related('participants'):
            prizes = influencer.get_prizes_list()
            participants_count = influencer.participants.count()
            
            ws.cell(row=row_num, column=1, value=influencer.id)
            ws.cell(row=row_num, column=2, value=influencer.name)
            # Get platform display - use model method or custom_platform
            if influencer.platform == 'other' and influencer.custom_platform:
                platform_display = influencer.custom_platform
            else:
                platform_display = influencer.get_platform_display()
            ws.cell(row=row_num, column=3, value=platform_display)
            ws.cell(row=row_num, column=4, value=influencer.username or '-')
            ws.cell(row=row_num, column=5, value=influencer.followers_count or 0)
            ws.cell(row=row_num, column=6, value=influencer.email or '-')
            ws.cell(row=row_num, column=7, value=influencer.phone or '-')
            ws.cell(row=row_num, column=8, value=influencer.get_status_display())
            ws.cell(row=row_num, column=9, value='نعم' if influencer.is_active else 'لا')
            ws.cell(row=row_num, column=10, value=len(prizes))
            ws.cell(row=row_num, column=11, value=participants_count)
            
            if influencer.created_at:
                ws.cell(row=row_num, column=12, value=influencer.created_at.strftime('%Y-%m-%d %H:%M'))
            else:
                ws.cell(row=row_num, column=12, value='-')
            
            if influencer.updated_at:
                ws.cell(row=row_num, column=13, value=influencer.updated_at.strftime('%Y-%m-%d %H:%M'))
            else:
                ws.cell(row=row_num, column=13, value='-')
            
            if influencer.approved_at:
                ws.cell(row=row_num, column=14, value=influencer.approved_at.strftime('%Y-%m-%d %H:%M'))
            else:
                ws.cell(row=row_num, column=14, value='-')
            
            row_num += 1
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if row.value:
                        max_length = max(max_length, len(str(row.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'المؤثرون_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save workbook to response
        wb.save(response)
        return response
    
    export_to_excel.short_description = "📊 تصدير البيانات المحددة إلى Excel"
    
    def approve_influencers(self, request, queryset):
        """Approve selected influencers"""
        count = 0
        for influencer in queryset:
            influencer.approve()
            count += 1
        self.message_user(request, f'تم الموافقة على {count} مؤثر')
    approve_influencers.short_description = 'الموافقة على المؤثرين المحددين'
    
    def reject_influencers(self, request, queryset):
        """Reject selected influencers"""
        count = queryset.update(status='rejected', is_active=False)
        self.message_user(request, f'تم رفض {count} مؤثر')
    reject_influencers.short_description = 'رفض المؤثرين المحددين'
    
    def activate_influencers(self, request, queryset):
        """Activate selected influencers"""
        count = queryset.update(is_active=True, status='active')
        self.message_user(request, f'تم تفعيل {count} مؤثر')
    activate_influencers.short_description = 'تفعيل المؤثرين المحددين'
    
    def deactivate_influencers(self, request, queryset):
        """Deactivate selected influencers"""
        count = queryset.update(is_active=False)
        self.message_user(request, f'تم إلغاء تفعيل {count} مؤثر')
    deactivate_influencers.short_description = 'إلغاء تفعيل المؤثرين المحددين'


@admin.register(Participant)
class ParticipantAdmin(admin.ModelAdmin):
    """Admin interface for Participant model"""
    
    list_display = [
        'id', 'name', 'phone', 'social_media_account', 'city', 
        'influencer', 'created_at_display'
    ]
    list_filter = ['influencer', 'city', 'created_at']
    search_fields = ['name', 'phone', 'social_media_account', 'city', 'influencer__name']
    readonly_fields = ['created_at']
    actions = ['export_to_excel']
    
    def changelist_view(self, request, extra_context=None):
        """Override changelist to add export button"""
        # Check if this is an export request
        if 'action' in request.POST and request.POST['action'] == 'export_to_excel':
            # Get all items (or selected items)
            selected_ids = request.POST.getlist('_selected_action')
            if selected_ids:
                queryset = self.get_queryset(request).filter(pk__in=selected_ids)
            else:
                # Export all if nothing selected
                queryset = self.get_queryset(request)
            return self.export_to_excel(request, queryset)
        
        extra_context = extra_context or {}
        extra_context['show_export_button'] = True
        extra_context['export_action_name'] = 'export_to_excel'
        return super().changelist_view(request, extra_context)
    
    fieldsets = (
        ('معلومات المشارك', {
            'fields': ('name', 'phone', 'social_media_account', 'city')
        }),
        ('معلومات المؤثر', {
            'fields': ('influencer',)
        }),
        ('التواريخ', {
            'fields': ('created_at',)
        }),
    )
    
    def created_at_display(self, obj):
        """Display created at in readable format"""
        if obj.created_at:
            return obj.created_at.strftime('%Y-%m-%d %H:%M')
        return '-'
    created_at_display.short_description = 'تاريخ التسجيل'
    
    def export_to_excel(self, request, queryset):
        """Export participants to Excel"""
        # Create a workbook and get the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "المشاركون"
        
        # Define headers in Arabic
        headers = [
            'ID',
            'الاسم',
            'رقم الجوال',
            'حساب التواصل الاجتماعي',
            'المدينة',
            'اسم المؤثر',
            'منصة المؤثر',
            'تاريخ التسجيل'
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="6A3FA0", end_color="6A3FA0", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        row_num = 2
        for participant in queryset.select_related('influencer'):
            ws.cell(row=row_num, column=1, value=participant.id)
            ws.cell(row=row_num, column=2, value=participant.name)
            ws.cell(row=row_num, column=3, value=participant.phone)
            ws.cell(row=row_num, column=4, value=participant.social_media_account)
            ws.cell(row=row_num, column=5, value=participant.city)
            ws.cell(row=row_num, column=6, value=participant.influencer.name)
            ws.cell(row=row_num, column=7, value=participant.influencer.get_platform_display())
            if participant.created_at:
                ws.cell(row=row_num, column=8, value=participant.created_at.strftime('%Y-%m-%d %H:%M'))
            else:
                ws.cell(row=row_num, column=8, value='-')
            row_num += 1
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if row.value:
                        max_length = max(max_length, len(str(row.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'المشاركون_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save workbook to response
        wb.save(response)
        return response
    
    export_to_excel.short_description = "📊 تصدير البيانات المحددة إلى Excel"
