"""
Views for influencers app
"""
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.views.generic import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
import random
import logging
from .models import Influencer, Participant

logger = logging.getLogger(__name__)


class InfluencerHomeView(TemplateView):
    """Home page view for influencers"""
    template_name = 'influencers/index.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'دوّرها | التسجيل للمؤثرين'
        return context


class InfluencerThanksView(TemplateView):
    """Thanks page view for influencers"""
    template_name = 'influencers/thanks.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        influencer_id = self.kwargs.get('influencer_id')
        if influencer_id:
            try:
                influencer = Influencer.objects.get(id=influencer_id)
                context['influencer'] = influencer
                context['influencer_url'] = influencer.influencer_url
            except Influencer.DoesNotExist:
                pass
        return context


@csrf_exempt
@require_http_methods(["POST"])
def register_influencer(request):
    """Register a new influencer"""
    try:
        data = json.loads(request.body)
        
        # Extract form data
        name = data.get('name', '').strip()
        platform = data.get('platform', '')
        custom_platform = data.get('customPlatform', '').strip()
        username = data.get('username', '').strip()
        profile_url = data.get('profile_url', '').strip()
        followers_count = data.get('followers_count', 0)
        email = data.get('email', '').strip()
        phone = data.get('phone', '').strip()
        
        # Process prizes (no percentages for influencers)
        prizes = data.get('prizes', [])
        
        # Validate required fields
        if not name:
            return JsonResponse({
                'success': False,
                'message': 'اسم المؤثر مطلوب'
            }, status=400)
        
        if not platform:
            return JsonResponse({
                'success': False,
                'message': 'المنصة مطلوبة'
            }, status=400)
        
        if not username:
            return JsonResponse({
                'success': False,
                'message': 'اسم المستخدم مطلوب'
            }, status=400)
        
        if not email:
            return JsonResponse({
                'success': False,
                'message': 'البريد الإلكتروني مطلوب'
            }, status=400)
        
        # If prizes is a string, convert it
        if isinstance(prizes, str):
            prizes = [prize.strip() for prize in prizes.split(',') if prize.strip()]
        
        # Validate prizes
        if not prizes or (isinstance(prizes, list) and len(prizes) == 0):
            return JsonResponse({
                'success': False,
                'message': 'يجب تحديد جوائز واحدة على الأقل'
            }, status=400)
        
        # Generate colors
        dawerha_colors = [
            "#6A3FA0", "#F2C23E", "#8C59C4", "#FF6B9D", 
            "#2E2240", "#4ECDC4", "#B794F6", "#FF8B5A"
        ]
        colors = random.sample(dawerha_colors, min(len(prizes), len(dawerha_colors)))
        
        # Validate followers_count
        try:
            followers_count = int(followers_count) if followers_count else 0
        except (ValueError, TypeError):
            followers_count = 0
        
        try:
            # Create influencer
            influencer = Influencer.objects.create(
                name=name,
                platform=platform,
                custom_platform=custom_platform if platform == 'other' else None,
                username=username,
                profile_url=profile_url if profile_url else None,
                followers_count=followers_count,
                email=email,
                phone=phone if phone else None,
                prizes=prizes,
                colors=colors,
                status='pending',
                is_active=False
            )
            
            from django.urls import reverse
            return JsonResponse({
                'success': True,
                'message': 'تم إرسال طلبك بنجاح',
                'influencer_id': influencer.id,
                'redirect_url': reverse('influencers:thanks', kwargs={'influencer_id': influencer.id})
            })
            
        except Exception as e:
            logger.error(f"Error creating influencer: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': f'حدث خطأ أثناء التسجيل: {str(e)}'
            }, status=500)
            
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'خطأ في البيانات المرسلة'
        }, status=400)
    except Exception as e:
        logger.error(f"Error in register_influencer: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        }, status=500)


def influencer_dashboard(request, influencer_id):
    """Influencer dashboard"""
    influencer = get_object_or_404(Influencer, id=influencer_id)
    
    # Get statistics
    participants_count = influencer.participants.count()
    
    context = {
        'influencer': influencer,
        'participants_count': participants_count,
    }
    
    return render(request, 'influencers/dashboard.html', context)


def export_participants_excel(request, influencer_id):
    """Export participants to Excel for a specific influencer"""
    influencer = get_object_or_404(Influencer, id=influencer_id)
    participants = influencer.participants.all()
    
    # Create a workbook and get the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "المشاركون"
    
    # Define headers in Arabic
    headers = [
        'ID',
        'الاسم',
        'رقم الجوال',
        'حساب التواصل الاجتماعي',
        'المدينة',
        'تاريخ التسجيل'
    ]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="6A3FA0", end_color="6A3FA0", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data
    row_num = 2
    for participant in participants:
        ws.cell(row=row_num, column=1, value=participant.id)
        ws.cell(row=row_num, column=2, value=participant.name)
        ws.cell(row=row_num, column=3, value=participant.phone)
        ws.cell(row=row_num, column=4, value=participant.social_media_account)
        ws.cell(row=row_num, column=5, value=participant.city)
        if participant.created_at:
            ws.cell(row=row_num, column=6, value=participant.created_at.strftime('%Y-%m-%d %H:%M'))
        else:
            ws.cell(row=row_num, column=6, value='-')
        row_num += 1
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for row in ws[column_letter]:
            try:
                if row.value:
                    max_length = max(max_length, len(str(row.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'المشاركون_{influencer.name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    return response


def register_participant_page(request, slug):
    """Page for participants to register"""
    influencer = get_object_or_404(Influencer, slug=slug)
    
    context = {
        'influencer': influencer,
    }
    
    return render(request, 'influencers/register_participant.html', context)


@csrf_exempt
@require_http_methods(["POST"])
def register_participant(request, slug):
    """Register a new participant"""
    try:
        influencer = get_object_or_404(Influencer, slug=slug)
        data = json.loads(request.body)
        
        name = data.get('name', '').strip()
        phone = data.get('phone', '').strip()
        social_media_account = data.get('social_media_account', '').strip()
        city = data.get('city', '').strip()
        
        # Validate required fields
        if not name:
            return JsonResponse({
                'success': False,
                'message': 'الاسم مطلوب'
            }, status=400)
        
        if not phone:
            return JsonResponse({
                'success': False,
                'message': 'رقم الجوال مطلوب'
            }, status=400)
        
        if not social_media_account:
            return JsonResponse({
                'success': False,
                'message': 'حساب التواصل الاجتماعي مطلوب'
            }, status=400)
        
        if not city:
            return JsonResponse({
                'success': False,
                'message': 'المدينة مطلوبة'
            }, status=400)
        
        # Create participant
        participant = Participant.objects.create(
            influencer=influencer,
            name=name,
            phone=phone,
            social_media_account=social_media_account,
            city=city
        )
        
        return JsonResponse({
            'success': True,
            'message': 'تم التسجيل بنجاح'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'خطأ في البيانات المرسلة'
        }, status=400)
    except Exception as e:
        logger.error(f"Error in register_participant: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        }, status=500)


def play_wheel_page(request, slug):
    """Wheel game page for influencer"""
    influencer = get_object_or_404(Influencer, slug=slug)
    
    prizes = influencer.get_prizes_list()
    colors = influencer.get_colors_list()
    
    # Get participants count
    participants_count = influencer.participants.count()
    
    context = {
        'influencer': influencer,
        'prizes': prizes,
        'colors': colors,
        'participants_count': participants_count,
    }
    
    return render(request, 'influencers/play_wheel.html', context)


@require_http_methods(["GET"])
def get_participants_count(request, slug):
    """Get current participants count"""
    try:
        influencer = get_object_or_404(Influencer, slug=slug)
        count = influencer.participants.count()
        return JsonResponse({
            'success': True,
            'count': count
        })
    except Exception as e:
        logger.error(f"Error getting participants count: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def spin_wheel(request, slug):
    """Handle wheel spin - select random prize and random winner"""
    try:
        influencer = get_object_or_404(Influencer, slug=slug)
        
        # Check if influencer is active
        if not influencer.is_active:
            return JsonResponse({
                'success': False,
                'message': 'المؤثر غير مفعل حالياً'
            }, status=403)
        
        # Get all participants
        participants = influencer.participants.all()
        
        if not participants.exists():
            return JsonResponse({
                'success': False,
                'message': 'لا يوجد مسجلين بعد'
            }, status=400)
        
        # Get prizes
        prizes = influencer.get_prizes_list()
        
        if not prizes:
            return JsonResponse({
                'success': False,
                'message': 'لا توجد جوائز متاحة'
            }, status=400)
        
        # Select random prize
        selected_prize = random.choice(prizes)
        
        # Select random participant (winner)
        winner = random.choice(list(participants))
        
        # Encrypt phone (hide last 4 digits, show first part)
        phone_encrypted = winner.phone
        if phone_encrypted and len(phone_encrypted) > 4:
            phone_encrypted = phone_encrypted[:-4] + '****'
        elif phone_encrypted and len(phone_encrypted) <= 4:
            phone_encrypted = '****'
        
        # Encrypt social media account (hide last 3 characters, show first part)
        social_encrypted = winner.social_media_account
        if social_encrypted and len(social_encrypted) > 3:
            social_encrypted = social_encrypted[:-3] + '***'
        elif social_encrypted and len(social_encrypted) <= 3:
            social_encrypted = '***'
        
        return JsonResponse({
            'success': True,
            'prize': selected_prize,
            'winner': {
                'name': winner.name,
                'phone': phone_encrypted,  # Encrypted for display
                'social_media_account': social_encrypted,  # Encrypted for display
                'city': winner.city
            }
        })
        
    except Exception as e:
        logger.error(f"Error in spin_wheel: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        }, status=500)
